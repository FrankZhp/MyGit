#! -*- coding utf-8 -*-
#! @Time  :2019/3/1 9:44
#! Author :Frank Zhang
#! @File  :Python_Blank.py
#! Python Version 3.7

#模块功能：读取某个文件夹下的所有FESTO Excel文件，读取每个文件里的所有Sheet，按照模板文件的样例，读取相应列，把数据合并到一个Excel文件中
#业务逻辑说明
#1.把模板文件的第一列读入list中,读取的FESTO文件，个别的可能与模板文件的列以及顺序不一样，那么按照模板文件的列及顺序，合并到一个新的文件中。
#2.模板文件里没有的列则不读取


from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
from tkinter.filedialog import askdirectory
import xlrd
import xlwt
import os
import openpyxl
import datetime
import time

Dict_ModExcelCol={}
modlist=[]

def main():
    #Get Directory
    def getDirectory():
        root.wm_attributes('-topmost',0)
        path_ = askdirectory()
        txt_Directory.insert(INSERT,path_)
        root.wm_attributes('-topmost',1)

    def selectExcelfile():
        root.wm_attributes('-topmost',0)
        sfname = filedialog.askopenfilename(title='Please Select Excel File', filetypes=[('Excel', '*.xlsx'), ('All Files', '*')])
        txt_ExcelModFile.insert(INSERT,sfname)
        root.wm_attributes('-topmost',1)
        
    def read_ModExcel():
        sModExcel=txt_ExcelModFile.get()
        wb = xlrd.open_workbook(filename=sModExcel)    #打开文件
        sheet1 = wb.sheet_by_index(0)                  #通过索引获取表格
        rows = sheet1.row_values(0)                    #获取行内容
        max_cols=sheet1.ncols

        #建立一个Dict，把列名和所在列的列号存入Dict中
        for i in range(max_cols):
            Dict_ModExcelCol[rows[i]]=i
            modlist.append(rows[i])
            
        #print(Dict_ModExcelCol)
        
    def closeThisWindow():
        root.destroy()

    def doProcess():
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示','开始合并FESTO Excel文件。')
        root.wm_attributes('-topmost',1)
        sPath=txt_Directory.get()
        #print(sPath)
        
        filenames = os.listdir(sPath)
        #print(filenames)
        
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(u'Sheet1')
        #为去重，定义一个list，遍历时，判断遍历到的TicketNo是否在此list中
        #如果不在，则增加到这个list中，并填写到结果文件中
        last_List=[]                            

        #iRow记录写的Excel的行号，第一行是从0开始，所以，初始化为-1
        iRow2=-1
        iCol2=-1

        #读取Excel模板文件，把模板文件里的的第一行的列名作为结果文件的列名
        read_ModExcel()

        #填写结果文件的第一行
        iRow2=0
        for i in range(len(Dict_ModExcelCol)):
            worksheet.write(iRow2,i,modlist[i])
     
        for iFile, filename in enumerate(filenames):
            #print('第:'+str(iFile+1)+'个文件： '+filename)
            sFileName=filename
                
            #wb = openpyxl.load_workbook(sPath+'\\'+sFileName)
            wb = xlrd.open_workbook(filename=sPath+'\\'+sFileName)
            # 获取workbook中所有的表格
            #sheets = wb.sheetnames
            sheets=wb.sheet_names()
            
            # 循环遍历所有sheet
            for i in range(len(sheets)):
                #sheet = wb[sheets[i]]
                sheet = wb.sheet_by_index(i)
                #print('第' + str(i + 1) + '个sheet Name: ' + sheet.title)
                #print('第' + str(i + 1) + '个sheet Name: ' + sheet.name)
                rows = sheet.row_values(0)   # 获取第一行内容
                cols = sheet.col_values(0)  #获取第1列的内容
                max_row=len(cols)
                max_column=len(rows)

                #判断这个Sheet的第一行第一个单元格是否是Ticket No.，如果不是，则不读取这一页
                sColumn1=sheet.cell(0,0).value
                #print("第一个单元格："+sColumn1)
                if sColumn1=="Ticket No":
                    pass
                else:
                    #print("退出这个sheet"+sheet.name+"File Name:"+sFileName)
                    continue
                   
                #print("Write Excel"+sheet.name)
                #第一列关键字，如果重复则去掉
                old_List=sheet.col_values(0)
                
                #第一行是列名，从第二行开始
                for iRow1 in range(1,max_row):
                   
                    for iCol1 in range(max_column):
                        if iCol1==0:
                            if old_List[iRow1] in last_List:                     #如果已有，则退出for循环，不增加重复数据
                                break                                   
                            else:
                                iRow2=iRow2+1
                                last_List.append(old_List[iRow1])                   #把没有关键字增加到列表中

                                #判断应该填写在哪一列，根据模板列来填写
                                iCol2=iCol1
                                worksheet.write(iRow2,iCol2,sheet.cell(iRow1,iCol1).value)
                        else:
                            #判断应该填写在哪一列，根据模板列来填写
                            if rows[iCol1] in Dict_ModExcelCol.keys():
                                iCol2=Dict_ModExcelCol.get(rows[iCol1])
                                worksheet.write(iRow2,iCol2,sheet.cell(iRow1,iCol1).value)

        strTime=time.strftime('%Y%m%d_%H%M%S',time.localtime(time.time())) 
        sPath=os.getcwd()+"\\Result\\"
        sReportName="2.FESTOReport_"+strTime+".xls"
        workbook.save(sPath+sReportName)

        print("Work is over.")
        root.wm_attributes('-topmost',0)
        tkinter.messagebox.showinfo('提示','处理完毕：\n'+sPath+sReportName)
        root.wm_attributes('-topmost',1)
    #初始化
    root=Tk()

    #设置窗体标题
    root.title('Merge FESTO files---Step 2')

    #设置窗口大小和位置
    root.geometry('660x300+430+220')

    label1=Label(root,text='Select FESTO Files Path:')
    txt_Directory=Entry(root,bg='white',width=62)
    btn_Browse1=Button(root,text='Browse',width=8,command=getDirectory)

    label2=Label(root,text='Select Mod File:',width=20,justify = tkinter.RIGHT)
    txt_ExcelModFile=Entry(root,bg='white',width=62)
    btn_Browse2=Button(root,text='Browse',width=8,command=selectExcelfile)
    
    btn_DoProcess=Button(root,text='Process',width=8,command=doProcess)
    btn_Exit=Button(root,text='Exit',width=8,command=closeThisWindow)
 

    label1.pack()
    txt_Directory.pack()
    label2.pack()
    txt_ExcelModFile.pack()

    btn_Browse1.pack()
    btn_Browse2.pack()
    
    btn_DoProcess.pack()
    btn_Exit.pack() 

    label1.place(x=30,y=30)
    txt_Directory.place(x=165,y=30)
    btn_Browse1.place(x=550,y=26)

    label2.place(x=30,y=60)
    txt_ExcelModFile.place(x=165,y=60)
    btn_Browse2.place(x=550,y=56)

    
    btn_DoProcess.place(x=230,y=100)
    btn_Exit.place(x=330,y=100)
 
 
    root.mainloop() 

def getFileNames(path):
    filenames = os.listdir(path)
    for i, filename in enumerate(filenames):
         if i==0:
            iSpecialFile=i+1
            sFileName=filename

def getSheetNames(path,sFileName):
    wb = openpyxl.load_workbook(path+'\\'+sFileName)
    # 获取workbook中所有的表格
    sheets = wb.sheetnames

    # 循环遍历所有sheet
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        print('第' + str(i + 1) + '个sheet Name: ' + sheet.title)
 
if __name__=="__main__":
    main()
